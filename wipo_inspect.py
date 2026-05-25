import zipfile, os, re
import openpyxl

# Check a file with MULTIPLE data rows
test_file = 'wipo_downloads/1165400_SWARNALI_SUNRISE_India.xlsx'
with zipfile.ZipFile(test_file, 'r') as z:
    names = z.namelist()
    media = [n for n in names if '/media/' in n and not n.endswith('/')]
    print(f'Media files ({len(media)}): {media}')

    wb = openpyxl.load_workbook(test_file)
    ws = wb['Export']
    print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')
    for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
        print(f'  {row}')

    if 'xl/drawings/drawing1.xml' in names:
        drw = z.read('xl/drawings/drawing1.xml').decode('utf-8')
        anchors = drw.count('oneCellAnchor')
        print(f'\nNumber of pic anchors: {anchors}')
        rows_found = re.findall(r'<xdr:row>(\d+)</xdr:row>', drw)
        cols_found = re.findall(r'<xdr:col>(\d+)</xdr:col>', drw)
        rids = re.findall(r'r:embed="(rId\d+)"', drw)
        print(f'Image row anchors (0-indexed): {rows_found}')
        print(f'Image col anchors (0-indexed): {cols_found}')
        print(f'rIds: {rids}')
        rels = z.read('xl/drawings/_rels/drawing1.xml.rels').decode('utf-8')
        print(f'Drawing rels:\n{rels}')
